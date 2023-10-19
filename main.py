import configparser
import re
import tkinter as tk
from datetime import datetime
from tkinter import filedialog

import matplotlib.pyplot as plt
import pandas as pd
import tabula
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

save_path = ''


def by_page(pdf_file_path, page):
    area = [35.0, 35.0, 735.0, 560.0]

    df_list = tabula.read_pdf(pdf_file_path,
                              multiple_tables=True,
                              pages=page,
                              area=area,
                              silent=True,  # Suppress all stderr output
                              )
    return df_list


def by_tabula(pdf_file_path):
    #  [top, left, bottom, right]
    area = [35.0, 35.0, 735.0, 560.0]
    # area = [80.0, 35.0, 735.0, 560.0]

    df_list = tabula.read_pdf(pdf_file_path,
                              multiple_tables=True,
                              pages="all",
                              area=area,
                              silent=True,  # Suppress all stderr output
                              )

    # 重新生成第一个
    area1 = [230.0, 35.0, 735.0, 560.0]
    df1 = tabula.read_pdf(pdf_file_path,
                          multiple_tables=True,
                          pages="1",
                          area=area1,
                          guess=False,
                          silent=True,  # Suppress all stderr output
                          )

    df_list[0] = df1[0]

    return df_list


def single_nan_cell_in_row(df, row_index):
    """
    判断DataFrame的指定行是否只有一个单元格包含nan。
    :param df: df (DataFrame): 要操作的DataFrame。
    :param row_index: 要检查的行的索引。
    :return: int: 如果只有一个单元格包含数据，返回列序号，否则返回-1。
    """

    # 不在数据区域直接返回 none
    if row_index <= 0 or row_index >= len(df):
        return -1

    # 使用loc来选择指定行
    row = df.loc[row_index]

    # 使用count()方法计算非NaN值的数量
    non_nan_count = row.count()

    if len(row) - non_nan_count == 1:
        for i, r in enumerate(row):
            if pd.isna(r):
                return i

    else:
        return -1


def is_single_data_cell_in_row(df, row_index):
    """
    判断DataFrame的指定行是否只有一个单元格包含数据，其他都是NaN。
    :param df: df (DataFrame): 要操作的DataFrame。
    :param row_index: 要检查的行的索引。
    :return: bool: 如果只有一个单元格包含数据，返回True，否则返回False。
    """

    # 不在数据区域直接返回 false
    if row_index <= 0 or row_index >= len(df) - 1:
        return False

    # 使用loc来选择指定行
    row = df.loc[row_index]

    # 使用count()方法计算非NaN值的数量
    non_nan_count = row.count()

    # 如果非NaN值的数量等于1，返回True；否则返回False
    return non_nan_count == 1


def has_num(text):
    pattern = re.compile(r'\d+')
    return bool(pattern.search(text))


def has_chinese(text):
    """
    检查文本中是否包含中文字符。

    参数：
    text (str): 要检查的文本。

    返回：
    bool: 如果文本中包含中文字符，返回True，否则返回False。
    """
    pattern = re.compile(r'[\u4e00-\u9fa5]')  # 中文字符的正则表达式范围
    return bool(pattern.search(text))


def insert_text_after_chinese(original_text, insert_text):
    # 使用正则表达式将字符串分割为中文、英文和数字，不分割日期
    pattern = r'([\u4e00-\u9fa5]+|[a-zA-Z]+|(?:\d{4}-\d{2}-\d{2})|\d+)'
    segments = re.findall(pattern, original_text)

    # 将插入文本插入在中文字符的后面
    new_text = ""
    for segment in segments:
        if re.match(r'[\u4e00-\u9fa5]+', segment):
            new_text += segment + insert_text
        else:
            new_text += segment

    return new_text


def remove_header_rows(df):
    """
    删除DataFrame中的表头行，用数字判断。
    :param df:  df (DataFrame): 要操作的DataFrame。
    :return: DataFrame: 删除没有中文字符的行后的新DataFrame。
    """

    # 使用apply和has_chinese函数检查每一行是否包含数字字符。出现第一个true后，后面不再判断全是true。
    mask = df.apply(lambda row: any(has_num(str(cell)) for cell in row), axis=1)
    i = 0
    for i, ma in enumerate(mask):
        if ma:
            break
    # 出现第一个true后，后面不再判断全是true。
    mask[i:] = True
    # 使用布尔索引选择行
    filtered_df = df[mask]

    # 重置索引
    filtered_df = filtered_df.reset_index(drop=True)

    return filtered_df


def data_clean(df_list):
    list_columns = ['记账日期', '货币', '交易金额', '联机余额', '交易摘要', '对手信息']
    for idx, df in enumerate(df_list):
        # 1.处理多行问题
        # 1.1 删除没有中文的一行
        df = remove_header_rows(df)
        i = 0
        # 1.2 遍历每行并处理最后一个单元格的 NaN 值
        while i < len(df):
            last_cell = df.iloc[i, -1]
            # 1.2.1 如果一行只有一个单元格的是 NaN ，并且它的上一行只有一个单元格有数据
            nan_id = single_nan_cell_in_row(df, i)
            if nan_id > -1 and is_single_data_cell_in_row(df, i - 1):
                # 最后一行特殊处理,增加一行假数据方便后面处理
                if i == len(df) - 1:
                    empty_row = pd.DataFrame([[''] * len(df.columns)], columns=df.columns)
                    # 使用 pd.concat 将空 DataFrame 与原始 DataFrame 连接
                    df = pd.concat([df, empty_row], ignore_index=True)
                # 合并上下两个单元格的数据到最后一个单元格
                df.iloc[i, nan_id] = str(df.iloc[i - 1, nan_id]) + str(df.iloc[i + 1, nan_id])
                # 删除上下两行的数据
                df = df.drop([i - 1, i + 1])
                df = df.reset_index(drop=True)
                i -= 2  # 因为删除了两行，所以回退两步
            # 1.2.2 如果最后一个单元格的是 NaN ，并且这行第一个单元格有中文，下一行只有一个单元格有数据
            if pd.isna(last_cell) and has_chinese(str(df.iloc[i, 0])) and is_single_data_cell_in_row(df, i + 1):
                # 把下一行最后一个单元格文本放到当前行第一个单元格的中文后面
                df.iloc[i, 0] = insert_text_after_chinese(str(df.iloc[i, 0]), str(df.iloc[i + 1, -1]))
                # 删除下一行，重置索引，因为删除了1行，所以回退1步
                df = df.drop([i + 1])
                df = df.reset_index(drop=True)
                i = i - 1
            i += 1

        # 重置索引
        df = df.reset_index(drop=True)
        df_list[idx] = df

        # 2.列数处理
        for column in df.columns:
            # 2.1 修改表头
            if column not in list_columns:
                # 以空格划分
                column_split = column.split(' ')
                for idx21, split in enumerate(column_split):
                    df.insert(idx21 + 1, 'new' + column_split[idx21], value='')
        df = df.reset_index(drop=True)
        df_list[idx] = df

        # 2.2 空白列(单元格)处理
        for column in df.columns:
            if column.startswith('new'):
                for idx22, cell in enumerate(df[column]):
                    if column == 'new记账日期':
                        # 使用正则表达式匹配日期（yyyy-mm-dd 格式）
                        date_pattern = r'\d{4}-\d{2}-\d{2}'  # 匹配 yyyy-mm-dd 格式的日期
                        dates = re.findall(date_pattern, df.iloc[idx22, 0])
                        df[column][idx22] = dates[0]
                        df.iloc[idx22, 0] = df.iloc[idx22, 0].replace(dates[0], '')
                    if column == 'new货币':
                        # 使用正则表达式匹配货币字符串
                        currency_pattern = r'[A-Z]{3}'  # 匹配三个大写字母作为货币代码
                        currencies = re.findall(currency_pattern, df.iloc[idx22, 0])
                        df[column][idx22] = currencies[0]
                        df.iloc[idx22, 0] = df.iloc[idx22, 0].replace(currencies[0], '')
            if column == '对手信息':
                for idx221, cell in enumerate(df[column]):
                    if pd.isna(cell):
                        df[column][idx221] = df.iloc[idx221, 0]
        df = df.rename(columns={'new记账日期': '记账日期'})
        df = df.rename(columns={'new货币': '货币'})
        df_list[idx] = df

        # 2.3 删除多余的列
        # 不在list_columns中的列
        items = [item for item in df.columns.tolist() if item not in list_columns]
        df.drop(axis=1, inplace=True, columns=items)

        print(f'第 {idx + 1}/{len(df_list)} 页数据处理完成...')

    # 3.合并数据list
    # 使用 concat 函数合并所有 DataFrame
    df = pd.concat(df_list, ignore_index=True)

    # 4.格式化数据
    # 设置日期列为datetime类型
    df["记账日期"] = pd.to_datetime(df["记账日期"])
    # 设置金额和余额为整数（分 0.01元）
    df["交易金额"] = (df["交易金额"].replace(",", "", regex=True).astype(float) * 100).astype(int)
    df["联机余额"] = (df["联机余额"].replace(",", "", regex=True).astype(float) * 100).astype(int)
    df["交易金额"] = pd.to_numeric(df["交易金额"])
    df["联机余额"] = pd.to_numeric(df["联机余额"])
    # 设置货币列为category类型
    df["货币"] = df["货币"].astype("category")
    # 设置str
    df["交易摘要"] = df["交易摘要"].astype("str")
    df["对手信息"] = df["对手信息"].astype("category")

    return df


def to_xlsx(df):
    # 创建一个 Excel 写入器，使用 with 语句来管理文件
    with pd.ExcelWriter(save_path + '/output.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)

        # 获取 Excel 的工作簿和工作表
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # 创建日期单元格格式
        date_format = 'yyyy-mm-dd'

        # 将各个格式应用到对于的单元格（假设数据从第二行开始）
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for col_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if isinstance(value, datetime):
                    cell.number_format = date_format
                if isinstance(value, int):
                    cell.value = cell.value / 100
                    cell.number_format = '0.00'

        # 对每列进行最合适的列宽操作
        for col_idx, column in enumerate(df, 1):
            max_length = max(df[column].astype(str).str.len().max(), len(column)) * 2
            worksheet.column_dimensions[get_column_letter(col_idx)].width = max_length


def format_number2k(value):
    # 将数字转换为"k"为单位的形式
    if abs(value) >= 1000:
        return f"{value / 1000:.1f}k"
    else:
        return f"{value:.1f}"


def to_pic(df):
    # 按照 "对手信息" 列进行分类统计 "交易金额" 列的总和
    data = df.groupby('对手信息', observed=True)['交易金额'].sum()
    # 将交易金额从 int64 转换为浮点数并保留两位小数
    result = (data / 100).astype('float').round(2)
    # 排序并选择前30个数据
    result = result.sort_values(ascending=False, key=abs).head(30)
    # 设置中文字体（示例使用微软雅黑，请根据您的系统和字体选择适合的字体）
    plt.rcParams['font.family'] = 'Microsoft YaHei'  # 使用微软雅黑
    plt.rcParams['axes.unicode_minus'] = False  # 解决负号显示为方块的问题
    plt.figure(figsize=(8, 4), dpi=500)
    plt.xticks(rotation=-45, fontsize=4)
    plt.yticks(rotation=45, fontsize=4)
    plt.xlabel('对手信息')
    plt.ylabel('交易金额')
    plt.title('对手信息 - 交易金额 （金额前30）')

    # 绘制柱状图
    ax = result.plot(kind='bar')
    # 在每个柱形上添加数据值
    for i, v in enumerate(result):
        # 调整数据值的位置，负数显示在图形下面
        va = 'top' if v < 0 else 'bottom'
        ax.text(i, v, format_number2k(v), ha='center', va=va, fontsize=4)
    # 保存柱状图为图片文件，设置tight bbox
    plt.savefig(save_path + '/bar_chart.png', bbox_inches='tight')


def open_file():
    print('请选择要打开的pdf文件：')
    # 创建 tkinter 窗口
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口
    # 打开文件对话框，等待用户选择文件
    file_path = filedialog.askopenfilename()
    # 检查用户是否选择了文件
    if file_path:
        print("文件路径是:", file_path)
    elif file_path == '':
        print("用户取消了文件选择")
    else:
        print("用户取消了文件选择")
    # 关闭 tkinter 窗口
    root.destroy()
    return file_path


def init_conf():
    global save_path

    # 创建配置解析器对象
    config = configparser.ConfigParser()
    # 读取配置文件
    config.read('config.ini')
    # 读取配置项

    save_path = config.get('App', 'save_path')
    print('已读取文件存储路径配置为： ' + save_path)


if __name__ == '__main__':
    print('欢迎使用招商银行pdf流水信息处理工具 v0.8.1')
    print('https://github.com/youzhiran')
    print('2668760098@qq.com')

    init_conf()

    pdf_file_path = open_file()

    print('1/4 数据读取...')
    df_list = by_tabula(pdf_file_path)
    # df_list = by_page(pdf_file_path, 56)

    # 数据清洗
    print('2/4 数据清洗...')
    df = data_clean(df_list)

    print('3/4 制作图表...')
    to_pic(df)

    print('4/4 制作表格...')
    to_xlsx(df)

    print('处理完成！')
    print('文件已存储在： ' + save_path)
